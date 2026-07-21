from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ENTETE_FILL = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
ENTETE_FONT = Font(color='FFFFFF', bold=True, size=11)
BORDURE = Border(*[Side(style='thin', color='CCCCCC')] * 4)


def _ecrire_entete(ws, colonnes, ligne=1):
    for idx, titre in enumerate(colonnes, start=1):
        cell = ws.cell(row=ligne, column=idx, value=titre)
        cell.fill = ENTETE_FILL
        cell.font = ENTETE_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDURE


def _ecrire_lignes(ws, lignes, ligne_depart=2):
    for i, ligne_data in enumerate(lignes, start=ligne_depart):
        for j, valeur in enumerate(ligne_data, start=1):
            cell = ws.cell(row=i, column=j, value=valeur)
            cell.border = BORDURE
            if i % 2 == 0:
                cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')


def _ajuster_largeurs(ws, colonnes):
    for idx, titre in enumerate(colonnes, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(len(titre) + 4, 16)


def generer_export_sites_excel(sites_queryset):
    """Export complet des sites, avec taux de survie calculé — pour analyse externe (Excel/Power BI)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sites de reboisement"

    colonnes = ['Nom', 'Localité', 'Province', 'Superficie (ha)', 'Statut',
                'Nombre de campagnes', 'Taux de survie moyen (%)', 'Responsable']
    _ecrire_entete(ws, colonnes)

    lignes = []
    for site in sites_queryset:
        lignes.append([
            site.nom, site.localite, site.province, float(site.superficie_hectares),
            site.get_statut_display(), site.campagnes.count(),
            site.taux_survie_moyen if site.taux_survie_moyen else 'N/A',
            site.responsable.get_full_name() if site.responsable else 'N/A',
        ])
    _ecrire_lignes(ws, lignes)
    _ajuster_largeurs(ws, colonnes)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generer_export_campagnes_excel(campagnes_queryset):
    """Export détaillé des campagnes, avec budget associé si présent."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Campagnes de plantation"

    colonnes = ['Site', 'Essence', 'Date plantation', 'Nombre plants',
                'Taux survie moyen (%)', 'Budget alloué', 'Coût réel', 'Responsable']
    _ecrire_entete(ws, colonnes)

    lignes = []
    for c in campagnes_queryset:
        budget = getattr(c, 'budget', None)
        lignes.append([
            c.site.nom, c.essence.nom, c.date_plantation.strftime('%d/%m/%Y'),
            c.nombre_plants,
            c.taux_survie_moyen if c.taux_survie_moyen else 'N/A',
            float(budget.budget_alloue) if budget else 'N/A',
            float(budget.cout_reel) if budget else 'N/A',
            c.responsable.get_full_name() if c.responsable else 'N/A',
        ])
    _ecrire_lignes(ws, lignes)
    _ajuster_largeurs(ws, colonnes)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generer_export_financements_excel(financements_queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Financements"

    colonnes = ['Partenaire', 'Type', 'Cible', 'Montant', 'Devise', 'Date', 'Référence']
    _ecrire_entete(ws, colonnes)

    lignes = []
    for f in financements_queryset:
        cible = f.campagne.__str__() if f.campagne else (f.site.nom if f.site else 'N/A')
        lignes.append([
            f.partenaire.nom, f.partenaire.get_type_partenaire_display(), cible,
            float(f.montant), f.devise, f.date_financement.strftime('%d/%m/%Y'), f.reference,
        ])
    _ecrire_lignes(ws, lignes)
    _ajuster_largeurs(ws, colonnes)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer